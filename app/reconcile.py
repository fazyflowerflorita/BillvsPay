from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


STANDARD_FIELDS = {
    "legal_entity",
    "employee_id",
    "personal_number",
    "employee_type",
    "pay_period",
    "earnings_code",
    "working_hours",
    "payroll_amount",
    "billing_amount",
    "customer",
    "invoice_number",
}

DEFAULT_COLUMN_ALIASES = {
    "legal entity": "legal_entity",
    "legalentity": "legal_entity",
    "company": "legal_entity",
    "entity": "legal_entity",
    "employee id": "employee_id",
    "employeeid": "employee_id",
    "payrollempid": "employee_id",
    "payroll emp id": "employee_id",
    "worker": "employee_id",
    "worker id": "employee_id",
    "personnel number": "personal_number",
    "personal number": "personal_number",
    "personalnumber": "personal_number",
    "personnel no": "personal_number",
    "person number": "personal_number",
    "employee type": "employee_type",
    "worker type": "employee_type",
    "employment type": "employee_type",
    "pay period": "pay_period",
    "period": "pay_period",
    "week ending": "pay_period",
    "transaction date": "pay_period",
    "week end": "pay_period",
    "earnings code": "earnings_code",
    "earning code": "earnings_code",
    "category id": "earnings_code",
    "earn code": "earnings_code",
    "pay code": "earnings_code",
    "working hours": "working_hours",
    "hours": "working_hours",
    "qty": "working_hours",
    "quantity": "working_hours",
    "payroll amount": "payroll_amount",
    "pay amount": "payroll_amount",
    "amount paid": "payroll_amount",
    "billing amount": "billing_amount",
    "bill amount": "billing_amount",
    "invoice amount": "billing_amount",
    "extended cost": "billing_amount",
    "customer": "customer",
    "customer name": "customer",
    "invoice": "invoice_number",
    "invoice number": "invoice_number",
}

MATCH_KEYS = ["employee_id", "earnings_code", "period_start", "period_end"]
TEXT_FIELDS = ["legal_entity", "employee_id", "personal_number", "employee_type", "earnings_code", "customer", "invoice_number"]
NUMBER_FIELDS = ["working_hours", "payroll_amount", "billing_amount"]
EARNINGS_CODE_ALIASES = {
    "OVERTIME": "OT",
    "EXPENSES": "EXPENSE",
    "EXP": "EXPENSE",
    "REGULAR TIME": "RT",
    "SICK": "SL",
    "SICK LEAVE": "SL",
    "SICK PAY": "SL",
}

EXPENSE_CODES = {"EXPENSE"}
SICK_LEAVE_CODES = {"SL"}
ROUNDING_TOLERANCE = 1.00

PAYROLL_OUTPUT_COLUMNS = [
    "Personnel number",
    "Worker",
    "Customer name",
    "Pay statement",
    "Payment date",
    "Earning code",
    "Description",
    "Begings Date",
    "Earnings date",
    "Quantity",
    "Rate",
    "Amount",
    "Pay cycle",
    "Pay period.Period start date",
    "Pay period.Period end date",
    "Position",
    "Posted",
    "HrsDifference",
    "AmntDifference",
    "Comments",
]

BILLING_OUTPUT_COLUMNS = [
    "PayrollEmpID",
    "SOP Number",
    "Invoice Date",
    "Item Number",
    "Category Id",
    "Item Description",
    "Placement ID",
    "Week Start",
    "Week End",
    "QTY",
    "Extended Cost",
    "Extended Price",
    "Unit Cost Price",
    "Unit Sales Price",
    "Customer Number",
    "Customer Name",
    "Batch Number",
    "SOP Type",
    "Payee Type",
    "Proj Contract Id",
    "Proj Id",
    "Document Amount",
    "GL Posting Date",
    "Created Date",
    "Posted Date",
    "MODIFDT",
    "Created By",
    "Posted By",
    "HrsDifference",
    "AmntDifference",
    "Comments",
]

STATUS_COLORS = {
    "Tallied": "C6EFCE",
    "Tallied - Consolidated Expense": "C6EFCE",
    "Tallied - Consolidated Sick Leave": "C6EFCE",
    "Tallied - Monthly Billing": "C6EFCE",
    "Rounded Off Difference": "E2F0D9",
    "Billed but not Paid": "FFC7CE",
    "Paid But not Billed": "F4B084",
    "Difference in working Hours": "FFEB9C",
    "Duplicate Invoice": "D9D2E9",
    "Over Billed": "FCE4D6",
    "Short Billed": "FCE4D6",
}

COLOR_LEGEND = [
    {"Status": "Tallied", "Color": "Green", "Meaning": "Payroll and Billing records matched within tolerance."},
    {"Status": "Tallied - Consolidated Expense", "Color": "Green", "Meaning": "Expense billing transactions were consolidated before matching to Payroll."},
    {"Status": "Tallied - Consolidated Sick Leave", "Color": "Green", "Meaning": "Split sick-leave Payroll/Billing rows were consolidated before matching."},
    {"Status": "Tallied - Monthly Billing", "Color": "Green", "Meaning": "Weekly Payroll records were consolidated to match monthly Billing."},
    {"Status": "Rounded Off Difference", "Color": "Light green", "Meaning": "Amount difference is below $1.00 and treated as rounding."},
    {"Status": "Billed but not Paid", "Color": "Red", "Meaning": "Billing record found without a matching Payroll record."},
    {"Status": "Paid But not Billed", "Color": "Orange", "Meaning": "Payroll record found without a matching Billing record."},
    {"Status": "Difference in working Hours", "Color": "Yellow", "Meaning": "Payroll and Billing matched, but hours differ."},
    {"Status": "Duplicate Invoice", "Color": "Purple", "Meaning": "More than one Billing record exists for the same match key."},
    {"Status": "Over Billed", "Color": "Light orange", "Meaning": "Billing Extended Cost is greater than Payroll Amount."},
    {"Status": "Short Billed", "Color": "Light orange", "Meaning": "Billing Extended Cost is less than Payroll Amount."},
]


@dataclass
class Rules:
    excluded_earnings_codes: set[str]
    included_earnings_codes: set[str]
    w2_only: bool = True
    hours_tolerance: float = 0.01
    amount_tolerance: float = 0.01
    column_aliases: dict[str, str] | None = None


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_code(value: Any) -> str:
    if pd.isna(value):
        return ""
    code = str(value).strip().upper()
    return EARNINGS_CODE_ALIASES.get(code, code)


def parse_date_series(series: pd.Series) -> pd.Series:
    dates = pd.to_datetime(series, errors="coerce")
    return dates.dt.strftime("%Y-%m-%d").fillna(series.astype("string").fillna("").str.strip())


def month_key_series(series: pd.Series) -> pd.Series:
    dates = pd.to_datetime(series, errors="coerce")
    return dates.dt.strftime("%Y-%m").fillna(series.astype("string").fillna("").str.slice(0, 7))


def period_span_days(start: pd.Series, end: pd.Series) -> pd.Series:
    start_dates = pd.to_datetime(start, errors="coerce")
    end_dates = pd.to_datetime(end, errors="coerce")
    return (end_dates - start_dates).dt.days.fillna(0)


def source_column(df: pd.DataFrame, *names: str) -> str | None:
    normalized = {normalize_header(column): column for column in df.columns}
    for name in names:
        found = normalized.get(normalize_header(name))
        if found is not None:
            return found
    return None


def clean_text_series(series: pd.Series) -> pd.Series:
    return series.astype("string").fillna("").str.strip()


def clean_money_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0.0).round(2)


def read_rules(path: str | Path) -> Rules:
    path = Path(path)
    aliases = DEFAULT_COLUMN_ALIASES.copy()
    excluded: set[str] = {"ACA", "BGC", "FEE"}
    included: set[str] = set()
    settings = {"w2_only": "true", "hours_tolerance": "0.01", "amount_tolerance": "0.01"}

    if path.exists():
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            workbook = pd.ExcelFile(path)
            if "column_mapping" in workbook.sheet_names:
                mapping_df = workbook.parse("column_mapping")
                for _, row in mapping_df.dropna(how="all").iterrows():
                    source = normalize_header(row.get("source_column"))
                    target = str(row.get("standard_field") or "").strip()
                    if source and target in STANDARD_FIELDS:
                        aliases[source] = target
            for sheet, target in [("excluded_earnings_codes", excluded), ("included_earnings_codes", included)]:
                if sheet in workbook.sheet_names:
                    values = workbook.parse(sheet).get("earnings_code", pd.Series(dtype=str))
                    target.update(normalize_code(v) for v in values if normalize_code(v))
            if "settings" in workbook.sheet_names:
                for _, row in workbook.parse("settings").dropna(how="all").iterrows():
                    key = str(row.get("setting") or "").strip()
                    value = str(row.get("value") or "").strip()
                    if key:
                        settings[key] = value
        else:
            rules_df = pd.read_csv(path)
            for _, row in rules_df.dropna(how="all").iterrows():
                rule_type = str(row.get("rule_type") or "").strip().lower()
                value = str(row.get("value") or "").strip()
                if rule_type == "excluded_earnings_code" and value:
                    excluded.add(normalize_code(value))
                elif rule_type == "included_earnings_code" and value:
                    included.add(normalize_code(value))
                elif rule_type == "setting" and value:
                    settings[str(row.get("description") or row.get("value") or "").strip()] = value

    return Rules(
        excluded_earnings_codes=excluded,
        included_earnings_codes=included,
        w2_only=str(settings.get("w2_only", "true")).lower() in {"true", "yes", "1"},
        hours_tolerance=float(settings.get("hours_tolerance", 0.01)),
        amount_tolerance=float(settings.get("amount_tolerance", 0.01)),
        column_aliases=aliases,
    )


def create_default_rules_workbook(path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    column_mapping = pd.DataFrame(
        [{"source_column": source, "standard_field": target} for source, target in sorted(DEFAULT_COLUMN_ALIASES.items())]
    )
    excluded = pd.DataFrame(
        [
            {"earnings_code": "ACA", "description": "Exclude ACA transactions"},
            {"earnings_code": "BGC", "description": "Exclude background check transactions"},
            {"earnings_code": "FEE", "description": "Exclude fee transactions"},
        ]
    )
    included = pd.DataFrame(columns=["earnings_code", "description"])
    settings = pd.DataFrame(
        [
            {"setting": "w2_only", "value": "true"},
            {"setting": "hours_tolerance", "value": 0.01},
            {"setting": "amount_tolerance", "value": 0.01},
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        column_mapping.to_excel(writer, sheet_name="column_mapping", index=False)
        excluded.to_excel(writer, sheet_name="excluded_earnings_codes", index=False)
        included.to_excel(writer, sheet_name="included_earnings_codes", index=False)
        settings.to_excel(writer, sheet_name="settings", index=False)


def load_report(path: str | Path, report_type: str, rules: Rules) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = pd.read_excel(path)
    original_count = len(raw)
    df = raw.dropna(how="all").copy()

    rename_map = {}
    aliases = rules.column_aliases or DEFAULT_COLUMN_ALIASES
    for column in df.columns:
        normalized = normalize_header(column)
        if normalized in aliases:
            rename_map[column] = aliases[normalized]
    df = df.rename(columns=rename_map)

    for field in list(STANDARD_FIELDS) + ["period_start", "period_end"]:
        if field not in df.columns:
            df[field] = pd.NA

    df = df[list(STANDARD_FIELDS.intersection(df.columns))].copy()
    for field in TEXT_FIELDS:
        if field in df.columns:
            df[field] = df[field].astype("string").fillna("").str.strip()

    if "earnings_code" in df.columns:
        df["earnings_code"] = df["earnings_code"].map(normalize_code)
    if "pay_period" in df.columns:
        df["pay_period"] = parse_date_series(df["pay_period"])
    for field in NUMBER_FIELDS:
        if field in df.columns:
            df[field] = pd.to_numeric(df[field], errors="coerce").fillna(0.0).round(2)

    before_dedupe = len(df)
    df = df.drop_duplicates().copy()

    if rules.w2_only and "employee_type" in df.columns:
        has_employee_type = df["employee_type"].astype(str).str.strip() != ""
        df = df[(~has_employee_type) | (df["employee_type"].str.upper() == "W2")].copy()

    if rules.included_earnings_codes:
        df = df[df["earnings_code"].isin(rules.included_earnings_codes)].copy()
    if rules.excluded_earnings_codes:
        df = df[~df["earnings_code"].isin(rules.excluded_earnings_codes)].copy()

    df["source_report"] = report_type
    audit = pd.DataFrame(
        [
            {"step": "Raw rows", "count": original_count},
            {"step": "After blank row removal", "count": len(raw.dropna(how="all"))},
            {"step": "Duplicate rows removed", "count": before_dedupe - len(raw.dropna(how="all").drop_duplicates())},
            {"step": "After filters and business rules", "count": len(df)},
        ]
    )
    return df.reset_index(drop=True), audit


def prepare_payroll(path: str | Path, rules: Rules) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = pd.read_excel(path).dropna(how="all").copy()
    df = raw.copy()
    df["_employee_id"] = clean_text_series(df[source_column(df, "Personnel number")]) if source_column(df, "Personnel number") else ""
    df["_earnings_code"] = df[source_column(df, "Earning code")].map(normalize_code) if source_column(df, "Earning code") else ""
    df["_period_start"] = parse_date_series(df[source_column(df, "Pay period.Period start date", "Begings Date")]) if source_column(df, "Pay period.Period start date", "Begings Date") else ""
    df["_period_end"] = parse_date_series(df[source_column(df, "Earnings date", "Pay period.Period end date")]) if source_column(df, "Earnings date", "Pay period.Period end date") else ""
    df["_period_month"] = month_key_series(df["_period_end"])
    df["_hours"] = clean_money_series(df[source_column(df, "Quantity")]) if source_column(df, "Quantity") else 0.0
    df["_amount"] = clean_money_series(df[source_column(df, "Amount")]) if source_column(df, "Amount") else 0.0

    payee_col = source_column(df, "Payee type name", "Payee Type")
    if rules.w2_only and payee_col:
        df = df[clean_text_series(df[payee_col]).str.upper().eq("W2")].copy()
    if rules.excluded_earnings_codes:
        df = df[~df["_earnings_code"].isin(rules.excluded_earnings_codes)].copy()
    if rules.included_earnings_codes:
        df = df[df["_earnings_code"].isin(rules.included_earnings_codes)].copy()
    df = df.drop_duplicates().reset_index(drop=True)

    if "Begings Date" not in df.columns:
        start_col = source_column(df, "Pay period.Period start date")
        df["Begings Date"] = df[start_col] if start_col else ""

    audit = pd.DataFrame(
        [
            {"step": "Raw rows", "count": len(raw)},
            {"step": "After W2 and earnings-code filters", "count": len(df)},
        ]
    )
    return df, audit


def prepare_billing(path: str | Path, rules: Rules) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = pd.read_excel(path).dropna(how="all").copy()
    df = raw.copy()
    df["_employee_id"] = clean_text_series(df[source_column(df, "PayrollEmpID", "Personalnumber")]) if source_column(df, "PayrollEmpID", "Personalnumber") else ""
    df["_earnings_code"] = df[source_column(df, "Category Id")].map(normalize_code) if source_column(df, "Category Id") else ""
    df["_period_start"] = parse_date_series(df[source_column(df, "Week Start")]) if source_column(df, "Week Start") else ""
    df["_period_end"] = parse_date_series(df[source_column(df, "Week End")]) if source_column(df, "Week End") else ""
    df["_period_month"] = month_key_series(df["_period_end"])
    df["_hours"] = clean_money_series(df[source_column(df, "QTY")]) if source_column(df, "QTY") else 0.0
    df["_amount"] = clean_money_series(df[source_column(df, "Extended Cost")]) if source_column(df, "Extended Cost") else 0.0

    payee_col = source_column(df, "Payee Type")
    if rules.w2_only and payee_col:
        df = df[clean_text_series(df[payee_col]).str.upper().eq("W2")].copy()
    if rules.excluded_earnings_codes:
        df = df[~df["_earnings_code"].isin(rules.excluded_earnings_codes)].copy()
    if rules.included_earnings_codes:
        df = df[df["_earnings_code"].isin(rules.included_earnings_codes)].copy()
    df = df.drop_duplicates().reset_index(drop=True)

    audit = pd.DataFrame(
        [
            {"step": "Raw rows", "count": len(raw)},
            {"step": "After W2 and earnings-code filters", "count": len(df)},
        ]
    )
    return df, audit


def aggregate_transactions(df: pd.DataFrame, amount_field: str) -> pd.DataFrame:
    agg = (
        df.groupby(MATCH_KEYS, dropna=False)
        .agg(
            working_hours=("working_hours", "sum"),
            amount=(amount_field, "sum"),
            transaction_count=("employee_id", "size"),
            personal_number=("personal_number", "first"),
            customer=("customer", "first"),
            invoice_number=("invoice_number", "first"),
        )
        .reset_index()
    )
    return agg


def aggregate_prepared(df: pd.DataFrame, prefix: str) -> pd.DataFrame:
    return (
        df.groupby(["_employee_id", "_earnings_code", "_period_start", "_period_end"], dropna=False)
        .agg(
            **{
                f"{prefix}_hours": ("_hours", "sum"),
                f"{prefix}_amount": ("_amount", "sum"),
                f"{prefix}_count": ("_employee_id", "size"),
            }
        )
        .reset_index()
    )


def is_consolidated_code(code: Any) -> bool:
    normalized = normalize_code(code)
    return normalized in EXPENSE_CODES or normalized in SICK_LEAVE_CODES


def match_scenario(code: Any, monthly_mode: bool) -> str:
    normalized = normalize_code(code)
    if normalized in EXPENSE_CODES:
        return "Expense Consolidation"
    if normalized in SICK_LEAVE_CODES:
        return "Sick Leave Consolidation"
    if monthly_mode:
        return "Monthly Billing"
    return "Weekly Exact"


def assign_match_keys(df: pd.DataFrame, monthly_mode: bool) -> pd.DataFrame:
    output = df.copy()
    use_month = output["_earnings_code"].map(is_consolidated_code) | monthly_mode
    output["_match_period"] = output["_period_start"].astype(str) + "|" + output["_period_end"].astype(str)
    output.loc[use_month, "_match_period"] = output.loc[use_month, "_period_month"]
    output["_match_scenario"] = output["_earnings_code"].map(lambda code: match_scenario(code, monthly_mode))
    return output


def aggregate_for_matching(df: pd.DataFrame, prefix: str, monthly_mode: bool) -> pd.DataFrame:
    keyed = assign_match_keys(df, monthly_mode)
    return (
        keyed.groupby(["_employee_id", "_earnings_code", "_match_period"], dropna=False)
        .agg(
            **{
                f"{prefix}_hours": ("_hours", "sum"),
                f"{prefix}_amount": ("_amount", "sum"),
                f"{prefix}_count": ("_employee_id", "size"),
            },
            period_start=("_period_start", "min"),
            period_end=("_period_end", "max"),
            period_month=("_period_month", "first"),
            match_scenario=("_match_scenario", "first"),
        )
        .reset_index()
    )


def detect_monthly_billing_mode(billing: pd.DataFrame) -> bool:
    if billing.empty:
        return False
    spans = period_span_days(billing["_period_start"], billing["_period_end"])
    monthly_rows = spans >= 27
    return bool(monthly_rows.mean() >= 0.25)


def build_ptl_payroll_output(payroll: pd.DataFrame, matched: pd.DataFrame, rules: Rules) -> pd.DataFrame:
    output = payroll.copy()
    monthly_mode = bool(matched.attrs.get("monthly_mode", False))
    output = assign_match_keys(output, monthly_mode)
    key_cols = ["_employee_id", "_earnings_code", "_match_period"]
    lookup = matched.set_index(key_cols)[["HrsDifference", "AmntDifference", "Comments"]].to_dict(orient="index")
    comments = []
    hour_diffs = []
    amount_diffs = []
    for _, row in output.iterrows():
        key = (row.get("_employee_id", ""), row.get("_earnings_code", ""), row.get("_match_period", ""))
        result = lookup.get(key, {})
        comment_value = result.get("Comments", "Paid But not Billed")
        comments.append(comment_value)
        hour_diff = result.get("HrsDifference", "")
        amount_diff = result.get("AmntDifference", "")
        hour_diffs.append("" if comment_value == "Tallied" or abs(float(hour_diff or 0)) <= rules.hours_tolerance else hour_diff)
        amount_diffs.append("" if comment_value == "Tallied" or abs(float(amount_diff or 0)) <= rules.amount_tolerance else amount_diff)
    output["Comments"] = comments
    output["HrsDifference"] = hour_diffs
    output["AmntDifference"] = amount_diffs
    for column in PAYROLL_OUTPUT_COLUMNS:
        if column not in output.columns:
            output[column] = ""
    return output[PAYROLL_OUTPUT_COLUMNS]


def build_ptl_billing_output(billing: pd.DataFrame, payroll: pd.DataFrame, rules: Rules) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    monthly_mode = detect_monthly_billing_mode(billing)
    payroll_agg = aggregate_for_matching(payroll, "payroll", monthly_mode)
    billing_agg = aggregate_for_matching(billing, "billing", monthly_mode)
    key_cols = ["_employee_id", "_earnings_code", "_match_period"]
    matched = billing_agg.merge(payroll_agg, on=key_cols, how="outer", indicator=True)
    for column in ["billing_hours", "billing_amount", "billing_count", "payroll_hours", "payroll_amount", "payroll_count"]:
        matched[column] = matched[column].fillna(0)
    for column in ["period_start", "period_end", "period_month", "match_scenario"]:
        left = f"{column}_x"
        right = f"{column}_y"
        if left in matched.columns and right in matched.columns:
            matched[column] = matched[left].fillna(matched[right])
            matched = matched.drop(columns=[left, right])
    matched["HrsDifference"] = (matched["payroll_hours"] - matched["billing_hours"]).round(2)
    matched["AmntDifference"] = (matched["payroll_amount"] - matched["billing_amount"]).round(2)

    def comment(row: pd.Series) -> str:
        scenario = str(row.get("match_scenario") or "")
        if row["_merge"] == "left_only":
            return "Billed but not Paid"
        if row["_merge"] == "right_only":
            return "Paid But not Billed"
        if row["billing_count"] > 1:
            if scenario == "Expense Consolidation" and abs(row["AmntDifference"]) <= rules.amount_tolerance:
                return "Tallied - Consolidated Expense"
            if scenario == "Expense Consolidation" and abs(row["AmntDifference"]) < ROUNDING_TOLERANCE:
                return "Rounded Off Difference"
            if scenario == "Sick Leave Consolidation" and abs(row["HrsDifference"]) <= rules.hours_tolerance and abs(row["AmntDifference"]) <= rules.amount_tolerance:
                return "Tallied - Consolidated Sick Leave"
            if scenario == "Monthly Billing" and abs(row["AmntDifference"]) <= rules.amount_tolerance:
                return "Tallied - Monthly Billing"
            return "Duplicate Invoice"
        if scenario in {"Expense Consolidation", "Monthly Billing"}:
            hours_mismatch = False
        else:
            hours_mismatch = abs(row["HrsDifference"]) > rules.hours_tolerance
        if hours_mismatch:
            return "Difference in working Hours"
        if abs(row["AmntDifference"]) > rules.amount_tolerance and abs(row["AmntDifference"]) < ROUNDING_TOLERANCE:
            return "Rounded Off Difference"
        if abs(row["AmntDifference"]) > rules.amount_tolerance:
            return "Over Billed" if row["AmntDifference"] < 0 else "Short Billed"
        if scenario == "Expense Consolidation":
            return "Tallied - Consolidated Expense"
        if scenario == "Sick Leave Consolidation":
            return "Tallied - Consolidated Sick Leave"
        if scenario == "Monthly Billing":
            return "Tallied - Monthly Billing"
        return "Tallied"

    matched["Comments"] = matched.apply(comment, axis=1)
    matched.attrs["monthly_mode"] = monthly_mode
    lookup = matched.set_index(key_cols)[["HrsDifference", "AmntDifference", "Comments"]].to_dict(orient="index")

    output = assign_match_keys(billing, monthly_mode)
    rename_pairs = {
        "Created By User Id": "Created By",
        "Posted By User Id": "Posted By",
    }
    output = output.rename(columns={source: target for source, target in rename_pairs.items() if source in output.columns})
    for column in BILLING_OUTPUT_COLUMNS:
        if column not in output.columns:
            output[column] = ""
    if "Personalnumber" in output.columns:
        output["PayrollEmpID"] = output["PayrollEmpID"].where(clean_text_series(output["PayrollEmpID"]) != "", output["Personalnumber"])

    comments = []
    hour_diffs = []
    amount_diffs = []
    for _, row in output.iterrows():
        key = (row.get("_employee_id", ""), row.get("_earnings_code", ""), row.get("_match_period", ""))
        result = lookup.get(key, {})
        comment_value = result.get("Comments", "")
        comments.append(comment_value)
        hour_diff = result.get("HrsDifference", "")
        amount_diff = result.get("AmntDifference", "")
        matched_like = comment_value.startswith("Tallied") or comment_value in {"Billed but not Paid", "Rounded Off Difference"}
        hour_diffs.append("" if matched_like or abs(float(hour_diff or 0)) <= rules.hours_tolerance else hour_diff)
        amount_diffs.append("" if comment_value.startswith("Tallied") or comment_value == "Billed but not Paid" or abs(float(amount_diff or 0)) <= rules.amount_tolerance else amount_diff)
    output["Comments"] = comments
    output["HrsDifference"] = hour_diffs
    output["AmntDifference"] = amount_diffs

    payroll_only = matched[matched["_merge"] == "right_only"].copy()
    if not payroll_only.empty:
        append_rows = []
        for _, row in payroll_only.iterrows():
            append_row = {column: "" for column in BILLING_OUTPUT_COLUMNS}
            append_row["PayrollEmpID"] = row["_employee_id"]
            append_row["Category Id"] = row["_earnings_code"]
            append_row["Week Start"] = row.get("period_start", "")
            append_row["Week End"] = row.get("period_end", "")
            append_row["QTY"] = row["payroll_hours"]
            append_row["Extended Cost"] = row["payroll_amount"]
            append_row["HrsDifference"] = row["HrsDifference"]
            append_row["AmntDifference"] = row["AmntDifference"]
            append_row["Comments"] = "Paid But not Billed"
            append_rows.append(append_row)
        output = pd.concat([output, pd.DataFrame(append_rows)], ignore_index=True)

    exception_summary = output["Comments"].value_counts(dropna=False).rename_axis("Comments").reset_index(name="Count")
    return output[BILLING_OUTPUT_COLUMNS], matched, exception_summary


def categorize(row: pd.Series, rules: Rules) -> str:
    if row["_merge"] == "left_only":
        return "Payroll Only"
    if row["_merge"] == "right_only":
        return "Billing Only"
    if row["payroll_transaction_count"] > 1 or row["billing_transaction_count"] > 1:
        return "Duplicate Transaction"
    if abs(row["hours_difference"]) > rules.hours_tolerance:
        return "Hours Mismatch"
    if abs(row["amount_difference"]) > rules.amount_tolerance:
        return "Amount Mismatch"
    return "Reconciled"


def apply_workbook_formatting(path: str | Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        if sheet.max_row < 1:
            continue
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        status_col = None
        for cell in sheet[1]:
            if str(cell.value or "").strip() in {"Comments", "Status"}:
                status_col = cell.column
                break

        if status_col:
            for row in range(2, sheet.max_row + 1):
                status = str(sheet.cell(row=row, column=status_col).value or "").strip()
                color = STATUS_COLORS.get(status)
                if not color:
                    continue
                fill = PatternFill("solid", fgColor=color)
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = fill

        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            max_length = 0
            for cell in sheet[letter][: min(sheet.max_row, 200)]:
                max_length = max(max_length, len(str(cell.value or "")))
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 34)

    if "Color Legend" in workbook.sheetnames:
        sheet = workbook["Color Legend"]
        status_col = 1
        for row in range(2, sheet.max_row + 1):
            status = str(sheet.cell(row=row, column=status_col).value or "").strip()
            color = STATUS_COLORS.get(status)
            if color:
                fill = PatternFill("solid", fgColor=color)
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = fill

    workbook.save(path)


def reconcile(payroll_path: str | Path, billing_path: str | Path, rules_path: str | Path, output_dir: str | Path) -> dict[str, Any]:
    rules = read_rules(rules_path)
    payroll, payroll_audit = prepare_payroll(payroll_path, rules)
    billing, billing_audit = prepare_billing(billing_path, rules)
    billing_output, matched, exception_summary = build_ptl_billing_output(billing, payroll, rules)
    payroll_output = build_ptl_payroll_output(payroll, matched, rules)

    reconciled_mask = matched["Comments"].astype(str).str.startswith("Tallied") | matched["Comments"].astype(str).eq("Rounded Off Difference")
    reconciled = matched[reconciled_mask].copy()
    exceptions = matched[~reconciled_mask].copy()
    summary = pd.DataFrame(
        [
            {"metric": "Payroll cleaned rows", "value": len(payroll)},
            {"metric": "Billing cleaned rows", "value": len(billing)},
            {"metric": "Reconciled transaction groups", "value": len(reconciled)},
            {"metric": "Exception transaction groups", "value": len(exceptions)},
            {"metric": "Payroll total amount", "value": round(float(payroll["_amount"].sum()), 2)},
            {"metric": "Billing total cost", "value": round(float(billing["_amount"].sum()), 2)},
            {"metric": "Amount difference", "value": round(float(payroll["_amount"].sum() - billing["_amount"].sum()), 2)},
        ]
    )

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"bill_vs_pay_reconciliation_{timestamp}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        payroll_output.to_excel(writer, sheet_name="Payroll Report", index=False)
        billing_output.to_excel(writer, sheet_name="Billing Report", index=False)
        exception_summary.to_excel(writer, sheet_name="Exception Summary", index=False)
        matched.to_excel(writer, sheet_name="Match Details", index=False)
        pd.DataFrame(COLOR_LEGEND).to_excel(writer, sheet_name="Color Legend", index=False)
        pd.concat([payroll_audit.assign(report="Payroll"), billing_audit.assign(report="Billing")]).to_excel(
            writer, sheet_name="Audit", index=False
        )
        pd.DataFrame(
            [{"rule": "Excluded Earnings Code", "value": code} for code in sorted(rules.excluded_earnings_codes)]
            + [{"rule": "Included Earnings Code", "value": code} for code in sorted(rules.included_earnings_codes)]
            + [
                {"rule": "W2 Only", "value": rules.w2_only},
                {"rule": "Hours Tolerance", "value": rules.hours_tolerance},
                {"rule": "Amount Tolerance", "value": rules.amount_tolerance},
                {"rule": "Rounding Tolerance", "value": f"Below ${ROUNDING_TOLERANCE:.2f}"},
                {"rule": "Expense Matching", "value": "Consolidate billing expense rows by employee, code, and month before matching amount."},
                {"rule": "Sick Leave Matching", "value": "Consolidate split sick-leave rows by employee, code, and month before matching hours and amount."},
                {"rule": "Monthly Billing Matching", "value": "When billing periods are monthly, consolidate weekly Payroll rows by employee, code, and month."},
            ]
        ).to_excel(writer, sheet_name="Business Rules", index=False)

    apply_workbook_formatting(output_path)

    return {
        "output_path": output_path,
        "summary": summary.to_dict(orient="records"),
        "exception_summary": exception_summary.to_dict(orient="records"),
        "reconciled_count": len(reconciled),
        "exception_count": len(exceptions),
    }
